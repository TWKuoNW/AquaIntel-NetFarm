import openpyxl
import os
from datetime import datetime

def find_log_folder_path():

    # get specified folder path
    log_folder_path = os.path.abspath(__file__)
    log_folder_path = os.path.dirname(log_folder_path)
    log_folder_path = log_folder_path.replace("\\", "/") + "/log/"

    return log_folder_path

def is_file_in_folder(folder_path, target_file_name):
    
    all_folder = os.listdir(folder_path) # get all folder

    # print(all_folder) # print all folder
    
    return (target_file_name in all_folder)

def create_a_new_excel(file_path_and_name):

    workbook = openpyxl.Workbook()
    # get default workbook
    sheet = workbook.active

    data = [
        ["時間", "溫度", "濕度"]
    ]
    for row in data:
        sheet.append(row)

    workbook.save(file_path_and_name)
    print("Excel created successfully")

def insert_data_into_excel(file_path_and_name, new_data):
    workbook = openpyxl.load_workbook(file_path_and_name)
    sheet = workbook.active
    next_row = sheet.max_row + 1
    for col_num, value in enumerate(new_data, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)
    workbook.save(file_path_and_name)

    print("Insertion successful")


log_folder_path = find_log_folder_path() # get log folder path
target_file_name = str(datetime.today().date()) + ".xlsx" # Use today's date as the filename

file_path_and_name = log_folder_path + target_file_name #

if is_file_in_folder(log_folder_path, target_file_name) == False:
    create_a_new_excel(file_path_and_name)


current_time = datetime.now().strftime("%H:%M:%S")
new_data = [current_time, 20, 60]
insert_data_into_excel(file_path_and_name, new_data)
