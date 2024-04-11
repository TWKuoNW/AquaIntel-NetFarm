"""
v2.2版本更新內容:
->優化程式碼架構
->新增MySQL的資料表
"""
import mysql.connector
from mysql.connector import Error
import time as t
from datetime import datetime
import threading
import openpyxl
import serial
import os
import tkinter as tk

# ================Modbus function block================
def modbusCRC(msg : str) -> int: # CRC calculator
    crc = 0xFFFF
    for n in range(len(msg)):
        crc ^= msg[n]
        for i in range(8):
            if(crc & 1):
                crc >>= 1
                crc ^= 0xA001
            else:
                crc >>= 1
    ba = crc.to_bytes(2, byteorder='little')
    return ba

def modbus_run(ser,origin_send,data_length):
    
    bytes_send = bytes([int(x, 16) for x in origin_send]) # list to bytes
    crc = modbusCRC(bytes_send) # CRC calculator function
    origin_send.append(str('{:02X}'.format(crc[0]))) # append crc LO 
    origin_send.append(str('{:02X}'.format(crc[1]))) # append crc HI     
    bytes_send = bytes([int(x, 16) for x in origin_send])

    ser.write(bytes_send) # send command
    data = ser.read(data_length) # read response

    return data
# ================Modbus function block================

# =================Excel function block=================
def creat_log_and_excel(excel_column):
    # get specified folder path and created log
    log_folder_path = os.path.abspath(__file__)
    log_folder_path = os.path.dirname(log_folder_path)
    log_folder_path = log_folder_path.replace("\\", "/") + "/log/"
    os.makedirs(log_folder_path, exist_ok=True)

    target_file_name = str(datetime.today().date()) + ".xlsx" # Use today's date as the filename
    file_path_and_name = log_folder_path + target_file_name  # log_folder_path + target_file_name

    all_document = os.listdir(log_folder_path) # Get log folder all document

    if (target_file_name in all_document) == False: # Creat excel
        workbook = openpyxl.Workbook()
        # get default workbook
        sheet = workbook.active

        for row in excel_column:
            sheet.append(row)

        workbook.save(file_path_and_name)
        print("Excel created successfully")
    
    return file_path_and_name

def insert_data_into_excel(file_path_and_name, new_data):
    workbook = openpyxl.load_workbook(file_path_and_name)
    sheet = workbook.active
    next_row = sheet.max_row + 1
    for col_num, value in enumerate(new_data, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)
    workbook.save(file_path_and_name)

    print("Excel insertion successful")
# =================Excel function block=================

# =================MySQL function block=================
def create_connection(host, port, user, password, database):
    connection = None
    try:
        connection = mysql.connector.connect(
            host = host,
            port = port,
            user = user,
            password = password,
            database = database
        )
        if connection.is_connected():
            print("Connected to MySQL database")
            return connection
    except Error as e:
        print(f"Error: {e}")
        return None

def create_table(connection, table_name, columns):
    try:
        cursor = connection.cursor()
        cursor.execute("SHOW TABLES")
        tables = cursor.fetchall()
        # Create the table if it hasn't been creat table today or database is empty.
        if(tables == []):
            create_table_query = f"CREATE TABLE IF NOT EXISTS {table_name} ({', '.join(columns)})"
            cursor.execute(create_table_query)
            connection.commit()
            print(f"Table {table_name} created successfully.")
        elif(not(table_name in tables[0])):
            create_table_query = f"CREATE TABLE IF NOT EXISTS {table_name} ({', '.join(columns)})"
            cursor.execute(create_table_query)
            connection.commit()
            print(f"Table {table_name} created successfully.")

        cursor.close()
    except Error as e:
        print(f"Error: {e}")

def insert_data_into_mysql(connection, table_name, data_dict):
    try:
        data_dict_keys = list(data_dict.keys()) # read keys

        cursor = connection.cursor()
        insert_command = "INSERT INTO " + table_name + " (" + data_dict_keys[0] + ", " + data_dict_keys[1] + ", " + data_dict_keys[2] + ") VALUES ( %s, %s, %s)"
        # print("Insert Command:", insert_command)
        cursor.execute(insert_command, (data_dict[data_dict_keys[0]], data_dict[data_dict_keys[1]], data_dict[data_dict_keys[2]]))
        connection.commit()
        print("MySQL inserted successfully")
    except Error as e:
        print(f"Error: {e}")
# =================MySQL function block=================

# ================Tkinter function block================
def write_to_text(data_dict):
    text_content = f"時間: {data_dict['time']}, 溫度: {data_dict['temperature']}, 濕度: {data_dict['humidity']}\n"
    text.insert(1.0, text_content) 
    
    text.tag_add("first_line", "1.0", "1.end")
    text.tag_add("other_line", "2.0", "end")
    text.tag_config("first_line", background="#BEBEBE")
    text.tag_config("other_line", background="white")

def close():
    global boolean
    boolean = False
    print("Exiting the program.")
    root.destroy()
    
def state():
    global check
    if(check):
        label_status.config(text="STOP", bg = "#FA8072")
        check = False
    else:
        label_status.config(text="RUN",  bg = "#02DF82")
        check = True
# ================Tkinter function block================
def main():
    # MySQL information
    host = "localhost"
    port = "3306"
    user = "root"
    password = "12345678"
    database = "sensordata"
    connection = create_connection(host, port, user, password, database)
    ser = serial.Serial('COM3', baudrate = 9600) # define COM PORT and baudrate
    # execute program
    if connection:
        try:
            while(boolean):   
                if(check):
                    origin_send = ["01", "04", "00", "01", "00", "02"] # request command
                    data = modbus_run(ser, origin_send, 9) # send command and get data
                    # ============= Conversion and translate =============
                    data = [format(x, '02x') for x in data]
                    value1 = data[3] + data[4]
                    value2 = data[5] + data[6]
                    temperature = int(value1, 16) # HEX to DEX
                    humidity = int(value2, 16) 

                    temperature = temperature / 10 # Temperature Conversion
                    humidity = humidity / 10 # Humidity Conversion

                    print("溫度:", temperature, " 濕度:", humidity)
                    # ============= Conversion and translate =============

                    day_and_current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S") 
                    current_time = datetime.now().strftime("%H:%M:%S")      

                    # ============= Save to Excel and MySQL =============
                    mysql_columns = [
                        "id INT AUTO_INCREMENT PRIMARY KEY",
                        "time DATETIME NOT NULL",
                        "temperature FLOAT NOT NULL",
                        "humidity FLOAT NOT NULL"
                    ]
                    original_date = str(datetime.today().date())
                    formatted_date = "t" + original_date.replace("-", "_")

                    create_table(connection, formatted_date, mysql_columns)
                    
                    data_column = [
                        ["時間", "溫度", "濕度"]
                    ]
                    data_row = [current_time, temperature, humidity]
                    
                    file_path_and_name = creat_log_and_excel(data_column)
                    
                    data_dict = { 
                        "time":day_and_current_time,
                        "temperature":temperature,
                        "humidity":humidity
                    }
                    insert_data_into_excel(file_path_and_name, data_row) # insert data to excel
                    insert_data_into_mysql(connection, formatted_date, data_dict) # data write to database
                    
                    write_to_text(data_dict) # data write to text
                    # ============= Save to Excel and MySQL =============                
                t.sleep(2)

        except Error as e:
            print("Error: ", e)

        finally:
            connection.close()
            print("Connection to MySQL database closed")    

main_thread = threading.Thread(target = main)  # define thread
main_thread.start() # thread start
boolean = True
check = True

# creat Tkinter
root = tk.Tk()
root.title("Sensor Reader")
root.protocol("WM_DELETE_WINDOW", close)

# creat label and button and Text Elements
label_status = tk.Label(root, text = "RUN", width=10, height=7, bg = "#02DF82" )
label_status.grid(row=0, column=0)
button_stop = tk.Button(root, text = "RUN/STOP", command = state, width=10, height=2)
button_stop.grid(row=1, column=0)
button_close = tk.Button(root, text = "CLOSE", command = close, width=10, height=2)
button_close.grid(row=2, column=0)
text = tk.Text(root, height = 20, width = 49)
text.grid(row=0, column=1 , rowspan=3)

# Run loop
root.mainloop()