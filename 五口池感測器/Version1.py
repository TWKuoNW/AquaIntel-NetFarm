import mysql.connector
from mysql.connector import Error
import time as t
from datetime import datetime
import threading
import openpyxl
import keyboard
import serial
import os

def modbusCRC(msg : str) -> int: # CRC calculator
    crc = 0xFFFF
    for n in range(len(msg)):
        crc ^= msg[n]
        for i in range(8):
            if(crc & 1):
                crc >>= 1
                crc ^= 0xA001
            else:
                crc >>= 1
    ba = crc.to_bytes(2, byteorder='little')
    return ba

# ================Excel function block================
def find_log_folder_path():

    # get specified folder path
    log_folder_path = os.path.abspath(__file__)
    log_folder_path = os.path.dirname(log_folder_path)
    log_folder_path = log_folder_path.replace("\\", "/") + "/log/"

    return log_folder_path

def is_file_in_folder(folder_path, target_file_name):
    
    all_folder = os.listdir(folder_path) # get all folder

    # print(all_folder) # print all folder
    
    return (target_file_name in all_folder)

def create_a_new_excel(file_path_and_name):

    workbook = openpyxl.Workbook()
    # get default workbook
    sheet = workbook.active

    data = [
        ["時間", "溫度", "濕度"]
    ]
    for row in data:
        sheet.append(row)

    workbook.save(file_path_and_name)
    print("Excel created successfully")

def insert_data_into_excel(file_path_and_name, new_data):
    workbook = openpyxl.load_workbook(file_path_and_name)
    sheet = workbook.active
    next_row = sheet.max_row + 1
    for col_num, value in enumerate(new_data, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)
    workbook.save(file_path_and_name)

    print("Excel insertion successful")
# ================Excel function block================

# ================MySQL function block================
def del_all_data(connection, table_name):
    cursor = connection.cursor()
    insert_command = "set SQL_SAFE_UPDATES = 0" # close safe function
    cursor.execute(insert_command)
    insert_command = "delete from sensordata." + "table_name" # clear datatable
    cursor.execute(insert_command)
    insert_command = "ALTER TABLE sensordata."+ table_name +" AUTO_INCREMENT = 1" # reset A.I.
    cursor.execute(insert_command)

def create_connection(host, port, user, password, database):
    connection = None
    try:
        connection = mysql.connector.connect(
            host = host,
            port = port,
            user = user,
            password = password,
            database = database
        )
        if connection.is_connected():
            print("Connected to MySQL database")
            return connection
    except Error as e:
        print(f"Error: {e}")
        return None

def create_table(connection, table_name, columns):
    try: 
        cursor = connection.cursor()
        create_table_query = f"CREATE TABLE {table_name} ({', '.join(columns)})"
        cursor.execute(create_table_query)
        connection.commit()
        cursor.close()
        print("Table ", table_name, " has been created.")
    except Error as e:
        print(f"Error: {e}")

def update_table(connection, table_name, new_columns):
    try: 
        cursor = connection.cursor()
        update_table_query = f"ALTER TABLE {table_name} ADD COLUMN {new_columns}"
        cursor.execute(update_table_query)
        connection.commit()
        cursor.close()
        print("Table ", table_name, " has been updated.")
    except Error as e:
        print(f"Error: {e}")

def drop_table(connection, table_name):
    try: 
        cursor = connection.cursor()
        drop_table_query = "DROP TABLE IF EXISTS " + table_name
        cursor.execute(drop_table_query)
        connection.commit()
        cursor.close()
        print("Table " , table_name, " has been dropped.")
    except Error as e:
        print(f"Error: {e}")

def show_table(connection):
    try:
        cursor = connection.cursor()
        cursor.execute("SHOW TABLES")
        tables = cursor.fetchall()
        for table  in tables:
            print(table)
        cursor.close()    
    except Error as e:
        print(f"Error: {e}")

def write_to_mysql(connection, table_name, data):
    try:
        cursor = connection.cursor()
        insert_command = "INSERT INTO " + table_name + " (time, temperature, humidity) VALUES ( %s, %s, %s)"
        # print("Insert Command:", insert_command)
        cursor.execute(insert_command, (data['time'], data['temperature'], data['humidity']))
        connection.commit()
        print("MySQL inserted successfully")
    except Error as e:
        print(f"Error: {e}")
# ================MySQL function block================
def modbus_run(ser):
    origin_send = ["01","04","00","01","00","02"] # request command
    bytes_send = bytes([int(x, 16) for x in origin_send]) # list to bytes

    crc = modbusCRC(bytes_send) # CRC calculator function
    origin_send.append(str('{:02X}'.format(crc[0]))) # append crc LO 
    origin_send.append(str('{:02X}'.format(crc[1]))) # append crc HI     
    
    # ==================== calc length ==================== 
    length = origin_send[4] + origin_send[5] 
    length = int(length , 16)  
    length = length * 2 + 5
    
    bytes_send = bytes([int(x, 16) for x in origin_send])
    # ==================== calc length ==================== 

    ser.write(bytes_send) # send command
    data = ser.read(length) # read response

    return data
 

def main():
    # MySQL information
    host = "localhost"
    port = "3306"
    user = "root"
    password = "12345678"
    database = "sensordata"
    connection = create_connection(host, port, user, password, database)
    ser = serial.Serial('COM4', baudrate = 9600) # define COM PORT and baudrate
    
    # execute program
    if connection:
        try:
            while(boolean):   
                data = modbus_run(ser)
                # ============= Conversion and translate =============
                data = [format(x, '02x') for x in data]
                value1 = data[3] + data[4]
                value2 = data[5] + data[6]
                temperature = int(value1, 16) # HEX to DEX
                humidity = int(value2, 16) 

                temperature = temperature / 10 # Temperature Conversion
                humidity = humidity / 10 # Humidity Conversion

                print("溫度:", temperature, " 濕度:", humidity)
                # ============= Conversion and translate =============
                

                # ============= Save to Excel =============
                log_folder_path = find_log_folder_path() # get log folder path
                target_file_name = str(datetime.today().date()) + ".xlsx" # Use today's date as the filename

                file_path_and_name = log_folder_path + target_file_name 

                if is_file_in_folder(log_folder_path, target_file_name) == False:
                    create_a_new_excel(file_path_and_name)

                current_time = datetime.now().strftime("%H:%M:%S")
                new_data = [current_time, temperature, humidity]
                insert_data_into_excel(file_path_and_name, new_data)
                # ============= Save to Excel =============
                time = datetime.now().strftime("%Y-%m-%d %H:%M:%S") 
                # set data
                data = { 
                    "time":time,
                    "temperature":temperature,
                    "humidity":humidity
                }
                write_to_mysql(connection, "data1", data) # data write to database
                
                t.sleep(10)

        except Error as e:
            print("Error: ", e)

        finally:
            connection.close()
            print("Connection to MySQL database closed")    

main_thread = threading.Thread(target = main)  # define thread
main_thread.start() # thread start
boolean = True

while(boolean):
    if keyboard.is_pressed('q'):
        if keyboard.is_pressed('w'):
            print("Exiting the program.")
            boolean = False

