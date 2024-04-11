"""
v2.1版本更新內容:
->優化程式碼架構
->UI功能優化，新增 "X" 退出功能
->移除非必要SQL Function
"""
import mysql.connector
from mysql.connector import Error
import time
from datetime import datetime
import threading
import openpyxl
import serial
import os
import tkinter as tk

# ================Modbus function block================
def modbusCRC(msg : str) -> int: # CRC calculator
    crc = 0xFFFF
    for n in range(len(msg)):
        crc ^= msg[n]
        for i in range(8):
            if(crc & 1):
                crc >>= 1
                crc ^= 0xA001
            else:
                crc >>= 1
    ba = crc.to_bytes(2, byteorder='little')
    return ba

def modbus_write(ser, origin_send):
    bytes_send = bytes([int(x, 16) for x in origin_send]) # list to bytes
    crc = modbusCRC(bytes_send) # CRC calculator function
    origin_send.append(str('{:02X}'.format(crc[0]))) # append crc LO 
    origin_send.append(str('{:02X}'.format(crc[1]))) # append crc HI     
    bytes_send = bytes([int(x, 16) for x in origin_send])
    ser.write(bytes_send) # send command

def modbus_read(ser, length):
    while(boolean):
        data = ser.read(length) # read response
        # ============= Conversion and translate =============
        data = [format(x, '02x') for x in data]
        print(data)
# ================Modbus function block================

# =================Excel function block=================
def creat_log_and_excel(excel_column):
    # get specified folder path and created log
    log_folder_path = os.path.abspath(__file__)
    log_folder_path = os.path.dirname(log_folder_path)
    log_folder_path = log_folder_path.replace("\\", "/") + "/log/"
    os.makedirs(log_folder_path, exist_ok=True)

    target_file_name = str(datetime.today().date()) + ".xlsx" # Use today's date as the filename
    file_path_and_name = log_folder_path + target_file_name  # log_folder_path + target_file_name

    all_document = os.listdir(log_folder_path) # Get log folder all document

    if (target_file_name in all_document) == False: # Creat excel
        workbook = openpyxl.Workbook()
        # get default workbook
        sheet = workbook.active

        for row in excel_column:
            sheet.append(row)

        workbook.save(file_path_and_name)
        print("Excel created successfully")
    
    return file_path_and_name

def insert_data_into_excel(file_path_and_name, new_data):
    workbook = openpyxl.load_workbook(file_path_and_name)
    sheet = workbook.active
    next_row = sheet.max_row + 1
    for col_num, value in enumerate(new_data, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)
    workbook.save(file_path_and_name)

    print("Excel insertion successful")
# =================Excel function block=================

# =================MySQL function block=================
def del_all_data(connection, table_name):
    cursor = connection.cursor()
    insert_command = "set SQL_SAFE_UPDATES = 0" # close safe function
    cursor.execute(insert_command)
    insert_command = "delete from sensordata." + "table_name" # clear datatable
    cursor.execute(insert_command)
    insert_command = "ALTER TABLE sensordata."+ table_name +" AUTO_INCREMENT = 1" # reset A.I.
    cursor.execute(insert_command)

def create_connection(host, port, user, password, database):
    connection = None
    try:
        connection = mysql.connector.connect(
            host = host,
            port = port,
            user = user,
            password = password,
            database = database
        )
        if connection.is_connected():
            print("Connected to MySQL database")
            return connection
    except Error as e:
        print(f"Error: {e}")
        return None

def write_to_mysql(connection, table_name, data):
    try:
        cursor = connection.cursor()
        insert_command = "INSERT INTO " + table_name + " (time, temperature, humidity) VALUES ( %s, %s, %s)"
        # print("Insert Command:", insert_command)
        cursor.execute(insert_command, (data['time'], data['temperature'], data['humidity']))
        connection.commit()
        print("MySQL inserted successfully")
    except Error as e:
        print(f"Error: {e}")
# =================MySQL function block=================

# ================Tkinter function block================
def write_to_text(data):
    text_content = f"時間: {data['time']}, 溫度: {data['temperature']}, 濕度: {data['humidity']}\n"
    text.insert(1.0, text_content) 
    
    text.tag_add("first_line", "1.0", "1.end")
    text.tag_add("other_line", "2.0", "end")
    text.tag_config("first_line", background="#BEBEBE")
    text.tag_config("other_line", background="white")

def close():
    global boolean
    boolean = False
    print("Exiting the program.")
    root.destroy()
    
def state():
    global check
    if(check):
        label_status.config(text="STOP", bg = "#FA8072")
        check = False
    else:
        label_status.config(text="RUN",  bg = "#02DF82")
        check = True

def LOG_window():
    log_window = tk.Toplevel(root)
    log_window.title("LOG")
    
    text_LOG = tk.Text(log_window, height = 20, width = 50)
    text_LOG.grid()
# ================Tkinter function block================
def main():
    ser = serial.Serial('COM3', baudrate = 9600) # define COM PORT and baudrate
    origin_send = ["01", "04", "00", "01", "00", "02"] # request command
    modbus_read_thread = threading.Thread(target = modbus_read, args = (ser, 9))  # define thread
    modbus_read_thread.start
    while(boolean):   
        if(check):
            modbus_write(ser, origin_send)
        time.sleep(2)


main_thread = threading.Thread(target = main)  # define thread
main_thread.start() # thread start
boolean = True
check = True

# creat Tkinter
root = tk.Tk()
root.title("Sensor Reader")
root.protocol("WM_DELETE_WINDOW", close)

# creat label and button and Text Elements
label_status = tk.Label(root, text = "RUN", width=10, height=7, bg = "#02DF82" )
label_status.grid(row=0, column=0)
button_stop = tk.Button(root, text = "RUN/STOP", command = state, width=10, height=2)
button_stop.grid(row=1, column=0)
button_close = tk.Button(root, text = "CLOSE", command = close, width=10, height=2)
button_close.grid(row=2, column=0)
button_LOG = tk.Button(root, text = "LOG_window", command = LOG_window, width=10, height=2)
button_LOG.grid(row=3, column=0)
text = tk.Text(root, height = 20, width = 49)
text.grid(row=0, column=1 , rowspan=4)

# Run loop
root.mainloop()